import pandas as pd
from tkinter import Tk, Label, Text, Button, END, messagebox, Entry
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

# Função para adicionar feedback e sugestão de melhoria
def add_feedback():
    global df  # Declarar df como global
    matricula = matricula_entry.get().strip()
    feedback = feedback_entry.get("1.0", END).strip()
    suggestion = suggestion_entry.get("1.0", END).strip()
    
    # Limitar o número de caracteres do feedback
    max_feedback_length = 500
    if len(feedback) > max_feedback_length:
        messagebox.showwarning("Feedback", f"O feedback deve ter no máximo {max_feedback_length} caracteres.")
        return
    
    if feedback:
        new_row = {
            'matricula': matricula if matricula else 'Não Informado',
            'date': datetime.now().strftime('%Y-%m-%d'),
            'feedback': feedback,
            'suggestion': suggestion
        }
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        # Usar um nome de arquivo temporário para evitar problemas de permissão
        temp_filename = 'feedbacks_temp.xlsx'
        df.to_excel(temp_filename, index=False)
        
        # Aplicar formatação ao arquivo Excel
        format_excel(temp_filename, 'feedbacks.xlsx')
        
        matricula_entry.delete(0, END)
        feedback_entry.delete("1.0", END)
        suggestion_entry.delete("1.0", END)
        messagebox.showinfo("Feedback", "Feedback e sugestão adicionados com sucesso!")
    else:
        messagebox.showwarning("Feedback", "Por favor, insira o feedback.")

def format_excel(input_filename, output_filename):
    # Carregar o arquivo temporário
    wb = load_workbook(input_filename)
    ws = wb.active

    # Definir estilos
    header_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(left=Side(border_style='thin'), 
                    right=Side(border_style='thin'), 
                    top=Side(border_style='thin'), 
                    bottom=Side(border_style='thin'))

    # Aplicar estilo ao cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    # Aplicar bordas e definir largura das colunas
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    # Ajustar largura das colunas
    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                column_widths[cell.column] = max((column_widths.get(cell.column, 0), len(str(cell.value))))

    for col, width in column_widths.items():
        ws.column_dimensions[chr(64 + col)].width = width + 2  # +2 para algum padding

    # Salvar o arquivo formatado
    wb.save(output_filename)

# Inicializar DataFrame
try:
    df = pd.read_excel('feedbacks.xlsx')
except FileNotFoundError:
    df = pd.DataFrame(columns=['matricula', 'date', 'feedback', 'suggestion'])

# Interface gráfica com tkinter
def on_enter(event):
    # Mudar o foco para o próximo campo
    if event.widget == feedback_entry:
        suggestion_entry.focus_set()

root = Tk()
root.title("Análise de Feedback dos Funcionários")

# Configurar cor de fundo da janela
root.configure(bg='#ADD8E6')  # Azul claro

# Configurar a grade
root.grid_rowconfigure(0, weight=1)
root.grid_rowconfigure(1, weight=1)
root.grid_rowconfigure(2, weight=1)
root.grid_rowconfigure(3, weight=1)
root.grid_rowconfigure(4, weight=1)
root.grid_rowconfigure(5, weight=1)
root.grid_rowconfigure(6, weight=1)
root.grid_columnconfigure(0, weight=1)

# Criar e posicionar widgets
Label(root, text="Digite seu número de matrícula (opcional):", bg='#ADD8E6').grid(row=0, column=0, sticky='w', padx=10, pady=5)
matricula_entry = Entry(root)
matricula_entry.grid(row=1, column=0, sticky='ew', padx=10, pady=5)

Label(root, text="Digite seu feedback (máximo de 500 caracteres):", bg='#ADD8E6').grid(row=2, column=0, sticky='w', padx=10, pady=5)
feedback_entry = Text(root, height=10, width=50)
feedback_entry.grid(row=3, column=0, sticky='ew', padx=10, pady=5)
feedback_entry.bind('<Return>', on_enter)  # Permitir usar Enter para ir para o próximo campo
feedback_entry.bind('<Tab>', on_enter)  # Permitir usar Tab para ir para o próximo campo

Label(root, text="Digite sua sugestão de melhoria (opcional):", bg='#ADD8E6').grid(row=4, column=0, sticky='w', padx=10, pady=5)
suggestion_entry = Text(root, height=5, width=50)
suggestion_entry.grid(row=5, column=0, sticky='ew', padx=10, pady=5)

Button(root, text="Adicionar Feedback e Sugestão", command=add_feedback).grid(row=6, column=0, pady=10)

root.mainloop()
